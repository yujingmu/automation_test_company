#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 12/19/2018 10:20 AM
# @Author  : yimi
# @File    : get_test_data.py

from openpyxl import load_workbook

class GetTestData:

    def __init__(self, file_path, mode, case_id_list):
        self.file_path = file_path
        self.mode = mode
        self.case_id_list = case_id_list

    def get_init_data(self):
        book = load_workbook(self.file_path)
        sheet = book["init_data"]
        init_data = sheet.cell(1, 2).value
        return init_data

    def update_init_data(self, new_number):
        book = load_workbook(self.file_path)
        sheet = book["init_data"]
        sheet.cell(1, 2).value = new_number
        book.save(self.file_path)

    def get_test_data(self, sheet_name):
        book = load_workbook(self.file_path)
        sheet = book[sheet_name]
        row = sheet.max_row
        data_list = []
        no_reg_tel = self.get_init_data()
        new_number = no_reg_tel+1
        # print(no_reg_tel)
        if self.mode == "1":
            for i in range(2, row+1):
                data_dict = {}
                case_id = sheet.cell(i, 1).value
                title = sheet.cell(i, 2).value
                url = sheet.cell(i, 3).value
                param = sheet.cell(i, 4).value
                method = sheet.cell(i, 5).value
                expected = sheet.cell(i, 6).value
                data_dict["case_id"] = case_id
                data_dict["title"] = title
                data_dict["url"] = url
                if param.find("${no_reg_tel}") != -1:
                    new_param = param.replace("${no_reg_tel}", str(no_reg_tel))
                    data_dict["param"] = new_param
                else:
                    data_dict["param"] = param

                data_dict["method"] = method
                data_dict["expected"] = expected
                data_list.append(data_dict)
        else:
            for i in eval(self.case_id_list):
                data_dict = {}
                case_id = sheet.cell(i+1, 1).value
                title = sheet.cell(i+1, 2).value
                url = sheet.cell(i+1, 3).value
                param = sheet.cell(i+1, 4).value
                method = sheet.cell(i+1, 5).value
                expected = sheet.cell(i+1, 6).value
                data_dict["case_id"] = case_id
                data_dict["title"] = title
                data_dict["url"] = url
                if param.find("${no_reg_tel}") != -1:
                    new_param = param.replace("${no_reg_tel}", str(no_reg_tel))
                    data_dict["param"] = new_param
                else:
                    data_dict["param"] = param

                data_dict["method"] = method
                data_dict["expected"] = expected
                data_list.append(data_dict)
        self.update_init_data(new_number)
        return data_list

    def write_data(self, r, c, sheet_name, new_data):
        book = load_workbook(self.file_path)
        sheet = book[sheet_name]
        sheet.cell(r, c).value = new_data
        book.save(self.file_path)

if __name__ == "__main__":
    test_data = GetTestData("test_data.xlsx").get_test_data("data")
    print(test_data)
    init_data = GetTestData("test_data.xlsx").get_init_data()
    # print(init_data)